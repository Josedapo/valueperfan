#!/usr/bin/env python3
"""
Process Excel data export from Horizm into ValuePerFan JSON files.

Usage:
  python3 scripts/process-excel.py <excel_file> [--month YYYY-MM] [--accounts-sheet NAME] [--countries-sheet NAME]

This script:
1. Reads the Excel file (accounts sheet + countries sheet)
2. Matches country IDs to country names
3. Normalizes categories
4. Validates and cleans data (duplicates, missing fields, VPF verification)
5. Calculates rankings per platform
6. Generates accounts.json, search-index.json
7. Updates history.json with the new month's ranking data
"""

import json
import os
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install openpyxl")
    sys.exit(1)


# --- Category mapping: raw Excel category → normalized display name ---
CATEGORY_MAP = {
    "Athlete": "Athlete",
    "Sport Teams": "Sport Team",
    "Sport Leagues": "Sport League",
    "Sport Federations": "Sport Organization",
    "Sport Events": "Sport Organization",
    "Sports Cups & Competitions": "Sport Organization",
    "Sports Foundations & Communities": "Sport Organization",
    "Music": "Musician",
    "Actor/Actress": "Actor",
    "Content Creators": "Content Creator",
    "Celebrity/Influencer": "Celebrity",
    "Media": "Media",
    "TV presenter": "Entertainment",
    "Comedian": "Entertainment",
    "Model": "Entertainment",
}

# --- Country name normalization: official long names → common short names ---
COUNTRY_NAME_MAP = {
    "Bolivia (Plurinational State of)": "Bolivia",
    "Congo (Democratic Republic of the)": "DR Congo",
    "Iran (Islamic Republic of)": "Iran",
    "Korea (Republic of)": "South Korea",
    "Korea (Democratic People's Republic of)": "North Korea",
    "Lao People's Democratic Republic": "Laos",
    "Macedonia (the former Yugoslav Republic of)": "North Macedonia",
    "Moldova (Republic of)": "Moldova",
    "Palestine, State of": "Palestine",
    "Russian Federation": "Russia",
    "Syrian Arab Republic": "Syria",
    "Tanzania, United Republic of": "Tanzania",
    "Venezuela (Bolivarian Republic of)": "Venezuela",
    "Viet Nam": "Vietnam",
    "Côte d'Ivoire": "Ivory Coast",
}

DEFAULT_AVATAR = "/images/default-avatar.svg"
SCRIPT_DIR = Path(__file__).parent
DATA_DIR = SCRIPT_DIR.parent / "src" / "data"


def build_country_map(ws):
    """Build {id: name} map from countries sheet."""
    countries = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            countries[int(row[0])] = row[2]
    return countries


def build_country_code_map(ws):
    """Build {id: code} map from countries sheet."""
    codes = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            codes[int(row[0])] = row[1]
    return codes


def normalize_category(raw_category):
    """Map raw Excel category to normalized category name."""
    if not raw_category:
        return None
    mapped = CATEGORY_MAP.get(raw_category.strip())
    if mapped:
        return mapped
    print(f"  WARNING: Unknown category '{raw_category}' — skipping account")
    return None


def process_excel(excel_path, month_label, accounts_sheet=None, countries_sheet=None):
    """Main processing pipeline."""
    print(f"Reading: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    # Select sheets by name or by index (default: first two)
    cs_name = countries_sheet or wb.sheetnames[1]
    as_name = accounts_sheet or wb.sheetnames[0]
    print(f"Accounts sheet: {as_name}")
    print(f"Countries sheet: {cs_name}")

    # Build country maps
    countries_ws = wb[cs_name]
    country_names = build_country_map(countries_ws)
    country_codes = build_country_code_map(countries_ws)

    # Read accounts
    accounts_ws = wb[as_name]
    raw_rows = list(accounts_ws.iter_rows(min_row=2, values_only=True))
    print(f"Raw rows: {len(raw_rows)}")

    # Stats
    stats = {
        "missing_critical": 0,
        "zero_value": 0,
        "duplicates": 0,
        "unknown_category": 0,
        "avatar_fallbacks": 0,
        "vpf_recalculated": 0,
        "country_missing": 0,
    }

    seen = set()
    accounts = []

    for row in raw_rows:
        # Unpack columns (A through O, pad short rows with None)
        padded = list(row) + [None] * (15 - len(row))
        (handle, platform, name, avatar_url, followers, posts,
         vpf, total_value, impressions, engagement, eng_rate,
         category, url, country_id, gender) = padded[:15]

        # Skip empty rows
        if not handle and not platform:
            continue

        # Critical fields check
        if not all([handle, platform, name, followers is not None, total_value is not None]):
            stats["missing_critical"] += 1
            continue

        # Normalize platform
        plat = str(platform).lower().strip()
        if plat not in ("instagram", "tiktok"):
            continue

        # Zero value check
        followers_int = int(followers) if followers else 0
        total_value_float = float(total_value) if total_value else 0.0
        if followers_int <= 0 or total_value_float <= 0:
            stats["zero_value"] += 1
            continue

        # Normalize handle (handle numeric values stored as floats by Excel)
        if isinstance(handle, float) and handle == int(handle):
            handle_clean = str(int(handle)).lower().strip()
        else:
            handle_clean = str(handle).lower().strip()

        # Duplicate check (same handle + platform)
        dup_key = f"{handle_clean}_{plat}"
        if dup_key in seen:
            stats["duplicates"] += 1
            continue
        seen.add(dup_key)

        # Category normalization
        normalized_cat = normalize_category(category)
        if normalized_cat is None:
            stats["unknown_category"] += 1
            continue

        # Recalculate VPF from total_value / followers (source of truth)
        calculated_vpf = total_value_float / followers_int
        if vpf is not None and not isinstance(vpf, str):
            if abs(calculated_vpf - float(vpf)) > 0.0001:
                stats["vpf_recalculated"] += 1

        # Avatar
        has_avatar = avatar_url and str(avatar_url).startswith("http")
        if not has_avatar:
            stats["avatar_fallbacks"] += 1

        # Country
        country_name = None
        country_code = None
        if country_id is not None:
            cid = int(country_id)
            country_name = country_names.get(cid)
            country_code = country_codes.get(cid)
            if country_name:
                country_name = COUNTRY_NAME_MAP.get(country_name, country_name)
            else:
                stats["country_missing"] += 1
        else:
            stats["country_missing"] += 1

        # Numeric fields (safe conversion)
        posts_int = int(posts) if posts else 0
        impressions_int = int(impressions) if impressions else 0
        engagement_int = int(engagement) if engagement else 0
        eng_rate_float = (engagement_int / impressions_int) if impressions_int > 0 else 0.0

        accounts.append({
            "handle": handle_clean,
            "platform": plat,
            "name": str(name).strip(),
            "avatarUrl": str(avatar_url).strip() if has_avatar else DEFAULT_AVATAR,
            "followers": followers_int,
            "posts": posts_int,
            "valuePerFan": round(calculated_vpf, 10),
            "totalValue": round(total_value_float, 2),
            "impressions": impressions_int,
            "engagement": engagement_int,
            "engRate": round(eng_rate_float, 6),
            "category": normalized_cat,
            "profileUrl": str(url).strip() if url else "",
            "country": country_name,
            "countryCode": country_code,
            "rank": {"vpf": 0, "totalValue": 0},
            "slug": handle_clean,
        })

    wb.close()

    # Calculate rankings per platform
    for plat in ("instagram", "tiktok"):
        plat_accounts = [a for a in accounts if a["platform"] == plat]

        # Rank by VPF (descending)
        by_vpf = sorted(plat_accounts, key=lambda a: a["valuePerFan"], reverse=True)
        for i, acc in enumerate(by_vpf):
            acc["rank"]["vpf"] = i + 1

        # Rank by Total Value (descending)
        by_tv = sorted(plat_accounts, key=lambda a: a["totalValue"], reverse=True)
        for i, acc in enumerate(by_tv):
            acc["rank"]["totalValue"] = i + 1

    # Build output
    ig_count = sum(1 for a in accounts if a["platform"] == "instagram")
    tk_count = sum(1 for a in accounts if a["platform"] == "tiktok")

    output = {
        "meta": {
            "lastUpdated": datetime.now().strftime("%Y-%m-%d"),
            "dataMonth": month_label,
            "totalAccounts": len(accounts),
            "platforms": {
                "instagram": ig_count,
                "tiktok": tk_count,
            },
        },
        "accounts": accounts,
    }

    # Build search index (lightweight)
    search_index = [
        {
            "handle": a["handle"],
            "name": a["name"],
            "platform": a["platform"],
            "slug": a["slug"],
            "avatarUrl": a["avatarUrl"],
            "category": a["category"],
            "country": a["country"],
        }
        for a in accounts
    ]

    # Write accounts.json
    accounts_path = DATA_DIR / "accounts.json"
    with open(accounts_path, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    # Write search-index.json (compact)
    search_path = DATA_DIR / "search-index.json"
    with open(search_path, "w", encoding="utf-8") as f:
        json.dump(search_index, f, ensure_ascii=False)

    # Update history
    update_history(accounts, month_label)

    # Report
    print(f"\n--- Processing Report ({month_label}) ---")
    print(f"Removed (missing critical fields): {stats['missing_critical']}")
    print(f"Removed (zero value/followers): {stats['zero_value']}")
    print(f"Removed (duplicates): {stats['duplicates']}")
    print(f"Removed (unknown category): {stats['unknown_category']}")
    print(f"Avatar fallbacks: {stats['avatar_fallbacks']}")
    print(f"VPF recalculated (mismatch): {stats['vpf_recalculated']}")
    print(f"Missing country: {stats['country_missing']}")
    print(f"\nFinal accounts: {len(accounts)}")
    print(f"  Instagram: {ig_count}")
    print(f"  TikTok: {tk_count}")

    # Category breakdown
    cats = {}
    for a in accounts:
        cats[a["category"]] = cats.get(a["category"], 0) + 1
    print(f"\nCategories:")
    for cat, count in sorted(cats.items(), key=lambda x: -x[1]):
        print(f"  {cat}: {count}")

    accounts_size = os.path.getsize(accounts_path) / 1024
    search_size = os.path.getsize(search_path) / 1024
    print(f"\nFiles written:")
    print(f"  {accounts_path} ({accounts_size:.0f} KB)")
    print(f"  {search_path} ({search_size:.0f} KB)")
    print(f"  {DATA_DIR / 'history.json'}")


def update_history(accounts, month_label):
    """
    Update history.json with ranking snapshots for the current month.

    Structure:
    {
      "months": ["2026-02", "2026-03", ...],
      "accounts": {
        "instagram:cristiano": {
          "2026-02": {"vpf": 1, "totalValue": 1},
          "2026-03": {"vpf": 2, "totalValue": 1}
        }
      }
    }
    """
    history_path = DATA_DIR / "history.json"

    if history_path.exists():
        with open(history_path, "r", encoding="utf-8") as f:
            history = json.load(f)
    else:
        history = {"months": [], "accounts": {}}

    # Add month if not already present
    if month_label not in history["months"]:
        history["months"].append(month_label)
        history["months"].sort()

    # Add ranking data for each account
    for acc in accounts:
        key = f"{acc['platform']}:{acc['handle']}"
        if key not in history["accounts"]:
            history["accounts"][key] = {}
        history["accounts"][key][month_label] = {
            "vpf": acc["rank"]["vpf"],
            "totalValue": acc["rank"]["totalValue"],
        }

    with open(history_path, "w", encoding="utf-8") as f:
        json.dump(history, f, ensure_ascii=False)

    print(f"\nHistory updated: {len(history['months'])} months tracked")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python3 process-excel.py <excel_file> [--month YYYY-MM] [--accounts-sheet NAME] [--countries-sheet NAME]")
        sys.exit(1)

    excel_file = sys.argv[1]

    def get_arg(flag):
        if flag in sys.argv:
            idx = sys.argv.index(flag)
            if idx + 1 < len(sys.argv):
                return sys.argv[idx + 1]
        return None

    month = get_arg("--month")
    accounts_sheet = get_arg("--accounts-sheet")
    countries_sheet = get_arg("--countries-sheet")

    if not month:
        month = "2026-02"  # Default for this first import
        print(f"No --month specified, using: {month}")

    process_excel(excel_file, month, accounts_sheet, countries_sheet)
